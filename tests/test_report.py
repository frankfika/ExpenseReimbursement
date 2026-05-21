"""Tests for app/report.py"""
import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from app.analyzer import InvoiceInfo
from app.report import ReportConfig, ReportGenerator, generate_report


def _make_invoice(**kwargs):
    """Helper: 创建带有默认值的 InvoiceInfo"""
    defaults = dict(
        type="餐费",
        subtype="",
        amount=100.0,
        date="2026-05-01",
        service_date="",
        merchant="Test Merchant",
        invoice_number="INV001",
        is_invoice=True,
        description="",
        raw_text="",
        file_path="/tmp/test.jpg",
    )
    defaults.update(kwargs)
    return InvoiceInfo(**defaults)


class TestReportConfig:
    """测试报表配置"""

    def test_default_values(self):
        config = ReportConfig()
        assert config.company_name == ""
        assert config.department == ""
        assert config.submitter == ""
        assert config.currency_symbol == "¥"
        assert config.include_charts is True
        assert config.date_format == "%Y-%m-%d"

    def test_custom_values(self):
        config = ReportConfig(
            company_name="TestCorp",
            department="Engineering",
            submitter="John Doe",
            currency_symbol="$",
            include_charts=False,
            date_format="%d/%m/%Y",
        )
        assert config.company_name == "TestCorp"
        assert config.department == "Engineering"
        assert config.submitter == "John Doe"
        assert config.currency_symbol == "$"
        assert config.include_charts is False
        assert config.date_format == "%d/%m/%Y"


class TestReportGenerator:
    """测试报表生成器"""

    def test_sanitize_sheet_name(self):
        generator = ReportGenerator("/tmp")
        assert generator._sanitize_sheet_name("正常名称") == "正常名称"
        assert generator._sanitize_sheet_name("名称:with:colons") == "名称_with_colons"
        assert generator._sanitize_sheet_name("名称/with/slashes") == "名称_with_slashes"
        assert generator._sanitize_sheet_name("名称\\with\\backslashes") == "名称_with_backslashes"
        assert generator._sanitize_sheet_name("名称?with?questions") == "名称_with_questions"
        assert generator._sanitize_sheet_name("名称*with*stars") == "名称_with_stars"
        assert generator._sanitize_sheet_name("名称[with]brackets") == "名称_with_brackets"
        assert generator._sanitize_sheet_name("a" * 50) == "a" * 31

    def test_generate_empty_categories(self):
        """空分类应生成空报表"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)
            result = generator.generate({})
            assert os.path.exists(result)
            assert Path(result).name == "报销统计.xlsx"

    def test_generate_with_invoices(self):
        """生成分类明细和汇总表"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            invoices = [
                _make_invoice(type="餐费", amount=100.50, merchant="Test Restaurant"),
                _make_invoice(type="打车票", amount=50.00, merchant="Didi"),
            ]

            categorized = {"餐费": [invoices[0]], "打车票": [invoices[1]]}
            result = generator.generate(categorized)

            # 验证文件存在
            assert os.path.exists(result)

            # 验证 Excel 内容
            wb = load_workbook(result)
            sheet_names = wb.sheetnames
            assert "汇总" in sheet_names

            # 检查汇总表
            summary = wb["汇总"]
            assert summary.cell(row=1, column=1).value is not None

            # 验证明细表存在
            assert any(name in sheet_names for name in ["餐费", "打车票"])

    def test_generate_skips_non_invoice(self):
        """只生成发票明细，跳过凭证"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            invoices = [
                _make_invoice(type="餐费", amount=100.50, is_invoice=True),
                _make_invoice(type="餐费", amount=0, is_invoice=False),
            ]

            categorized = {"餐费": invoices}
            result = generator.generate(categorized)

            wb = load_workbook(result)
            # 明细表应该只有发票数据（1行数据 + 表头 + 合计）
            ws = None
            for name in wb.sheetnames:
                if name != "汇总":
                    ws = wb[name]
                    break

            assert ws is not None
            # 应该有表头(第1行)、数据(第2行)、合计(第3行)
            assert ws.cell(row=2, column=1).value == 1  # 序号1

    def test_generate_with_company_info(self):
        """包含公司信息的报表"""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ReportConfig(
                company_name="TestCorp Inc.",
                department="R&D",
                submitter="Jane Smith",
            )
            generator = ReportGenerator(tmpdir, config=config)

            invoices = [_make_invoice(type="办公用品", amount=200.00, merchant="Office Depot")]
            categorized = {"办公用品": invoices}
            result = generator.generate(categorized)

            wb = load_workbook(result)
            summary = wb["汇总"]
            # 公司名应在第一行
            assert summary.cell(row=1, column=1).value == "TestCorp Inc."

    def test_generate_sorts_by_date(self):
        """明细表应按日期排序"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            invoices = [
                _make_invoice(date="2026-05-03", merchant="Later"),
                _make_invoice(date="2026-05-01", merchant="Earlier"),
            ]

            categorized = {"餐费": invoices}
            result = generator.generate(categorized)

            wb = load_workbook(result)
            ws = None
            for name in wb.sheetnames:
                if name != "汇总":
                    ws = wb[name]
                    break

            assert ws is not None
            # 按日期排序后，5-01 应该在前面
            assert ws.cell(row=2, column=2).value == "2026-05-01"
            assert ws.cell(row=3, column=2).value == "2026-05-03"

    def test_generate_with_relative_path(self):
        """文件路径应显示为相对路径"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            # 使用与 output_dir 相关的路径
            rel_file = Path(tmpdir) / "invoices" / "test.jpg"
            rel_file.parent.mkdir(parents=True, exist_ok=True)
            rel_file.touch()

            invoices = [_make_invoice(file_path=str(rel_file))]
            categorized = {"餐费": invoices}
            result = generator.generate(categorized)

            wb = load_workbook(result)
            ws = None
            for name in wb.sheetnames:
                if name != "汇总":
                    ws = wb[name]
                    break

            assert ws is not None
            path_value = ws.cell(row=2, column=7).value
            assert "invoices" in str(path_value)

    def test_detail_sheet_sum_formula(self):
        """明细表合计行应包含 SUM 公式"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            invoices = [_make_invoice()]
            categorized = {"餐费": invoices}
            result = generator.generate(categorized)

            wb = load_workbook(result)
            ws = None
            for name in wb.sheetnames:
                if name != "汇总":
                    ws = wb[name]
                    break

            assert ws is not None
            # 合计行在第3行
            sum_formula = ws.cell(row=3, column=4).value
            assert "=SUM(" in str(sum_formula)

    def test_summary_formulas(self):
        """汇总表应包含引用明细表的公式"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            invoices = [_make_invoice()]
            categorized = {"餐费": invoices}
            result = generator.generate(categorized)

            wb = load_workbook(result)
            summary = wb["汇总"]

            # 找到数据行：先找到表头"金额（元）"所在行，数据在其下一行
            header_row = None
            for row in range(1, 20):
                if summary.cell(row=row, column=3).value == "金额（元）":
                    header_row = row
                    break

            assert header_row is not None, "未找到表头"
            data_row = header_row + 1
            amount_formula = summary.cell(row=data_row, column=3).value
            assert "=" in str(amount_formula), f"Expected formula, got: {amount_formula}"

    def test_category_order(self):
        """类别应按预定义顺序排列"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            categorized = {
                "其他": [_make_invoice(type="其他", amount=10.00, merchant="Other")],
                "餐费": [_make_invoice(type="餐费", amount=50.00, merchant="Restaurant")],
                "打车票": [_make_invoice(type="打车票", amount=30.00, merchant="Taxi")],
            }

            result = generator.generate(categorized)
            wb = load_workbook(result)
            summary = wb["汇总"]

            # 找到数据行
            categories_in_sheet = []
            for row in range(1, 20):
                val = summary.cell(row=row, column=1).value
                if val in ["打车票", "餐费", "其他"]:
                    categories_in_sheet.append(val)

            # 打车票应该在餐费之前，餐费在其他之前
            assert categories_in_sheet.index("打车票") < categories_in_sheet.index("餐费")
            assert categories_in_sheet.index("餐费") < categories_in_sheet.index("其他")

    def test_add_pie_chart(self):
        """测试添加饼图"""
        from openpyxl.chart.pie_chart import PieChart

        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            invoices = [_make_invoice()]
            categorized = {"餐费": invoices}
            result = generator.generate(categorized)

            wb = load_workbook(result)
            summary = wb["汇总"]
            # 验证图表已添加到工作表
            assert len(summary._charts) == 1
            chart = summary._charts[0]
            assert isinstance(chart, PieChart)
            # 验证标题
            assert chart.title.text.rich.p[0].r[0].t == "报销金额分布"

    def test_empty_category_with_no_invoices(self):
        """没有发票的类别不应该创建工作表"""
        with tempfile.TemporaryDirectory() as tmpdir:
            generator = ReportGenerator(tmpdir)

            categorized = {
                "餐费": [],  # 空列表
                "打车票": [_make_invoice(type="打车票", amount=30.00, merchant="Taxi")],
            }

            result = generator.generate(categorized)
            wb = load_workbook(result)
            # 只有汇总和打车票两个表
            assert len(wb.sheetnames) == 2


class TestGenerateReport:
    """测试便捷函数"""

    def test_generate_report(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            invoices = [_make_invoice()]

            categorized = {"餐费": invoices}
            result = generate_report(
                tmpdir,
                categorized,
                company_name="TestCorp",
                department="Dev",
                submitter="John",
            )

            assert os.path.exists(result)
            wb = load_workbook(result)
            summary = wb["汇总"]
            assert summary.cell(row=1, column=1).value == "TestCorp"
