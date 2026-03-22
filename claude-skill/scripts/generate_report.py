#!/usr/bin/env python3
"""
报销统计报表生成器

用法:
  # 单周期报表
  python generate_report.py <output_dir> <invoice_data.json>

  # 多周期总汇总报表
  python generate_report.py <output_dir> <periods_summary.json> --summary
"""
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 未安装，正在安装...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter


# Shared styles
def get_styles():
    return {
        "header_font_white": Font(bold=True, size=11, color="FFFFFF"),
        "header_font": Font(bold=True, size=11),
        "title_font": Font(bold=True, size=16),
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "period_fill": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "border": Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
        "center": Alignment(horizontal='center', vertical='center'),
        "right": Alignment(horizontal='right', vertical='center'),
        "wrap": Alignment(horizontal='left', vertical='center', wrap_text=True),
    }


def sanitize_sheet_name(name: str) -> str:
    for char in [':', '/', '\\', '?', '*', '[', ']']:
        name = name.replace(char, '_')
    return name[:31]


def generate_report(output_dir: str, invoices: list) -> str:
    """生成单周期报销报表"""
    output_path = Path(output_dir)
    wb = Workbook()
    wb.remove(wb.active)
    s = get_styles()

    category_order = ['打车票', '火车飞机票', '住宿费', '餐费', '其他']

    # Group by category - only invoices (is_invoice=true)
    categories = {}
    for inv in invoices:
        if not inv.get('is_invoice', True):
            continue
        cat = inv.get('category_name', '其他')
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(inv)

    # Create detail sheets
    sheet_info = {}
    for cat_name in category_order:
        if cat_name not in categories or not categories[cat_name]:
            continue
        items = sorted(categories[cat_name], key=lambda x: x.get('date', '') or '')
        sheet_name = sanitize_sheet_name(cat_name)
        ws = wb.create_sheet(sheet_name)

        headers = ['序号', '日期', '商家/平台', '金额（元）', '发票号码', '描述']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = s["header_font_white"]
            cell.fill = s["header_fill"]
            cell.border = s["border"]
            cell.alignment = s["center"]

        for idx, inv in enumerate(items, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx).border = s["border"]
            ws.cell(row=row, column=1).alignment = s["center"]

            date_val = inv.get('service_date') or inv.get('date', '')
            ws.cell(row=row, column=2, value=date_val).border = s["border"]
            ws.cell(row=row, column=2).alignment = s["center"]

            merchant = inv.get('subtype') or inv.get('merchant', '')
            ws.cell(row=row, column=3, value=merchant).border = s["border"]

            cell = ws.cell(row=row, column=4, value=inv.get('amount', 0))
            cell.border = s["border"]
            cell.alignment = s["right"]
            cell.number_format = '#,##0.00'

            ws.cell(row=row, column=5, value=inv.get('invoice_number', '')).border = s["border"]
            ws.cell(row=row, column=6, value=inv.get('description', '')).border = s["border"]
            ws.cell(row=row, column=6).alignment = s["wrap"]

        # Sum row
        sum_row = len(items) + 2
        ws.cell(row=sum_row, column=3, value="合计").font = s["header_font"]
        ws.cell(row=sum_row, column=3).border = s["border"]
        cell = ws.cell(row=sum_row, column=4, value=f'=SUM(D2:D{sum_row - 1})')
        cell.font = s["header_font"]
        cell.border = s["border"]
        cell.alignment = s["right"]
        cell.number_format = '#,##0.00'
        for col in [1, 2, 5, 6]:
            ws.cell(row=sum_row, column=col, value="").border = s["border"]

        col_widths = [6, 12, 20, 12, 20, 30]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        sheet_info[cat_name] = (sheet_name, len(items))

    # Summary sheet
    ws = wb.create_sheet("汇总", 0)
    ws.merge_cells('A1:D1')
    ws['A1'] = f"报销汇总表 - {datetime.now().strftime('%Y-%m-%d')}"
    ws['A1'].font = s["title_font"]
    ws['A1'].alignment = s["center"]

    headers = ['类别', '发票数量', '金额（元）', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = s["header_font_white"]
        cell.fill = s["header_fill"]
        cell.border = s["border"]
        cell.alignment = s["center"]

    row = 4
    count_cells = []
    amount_cells = []

    for cat_name in category_order:
        if cat_name not in sheet_info:
            continue
        sname, count = sheet_info[cat_name]
        last_data_row = count + 1

        ws.cell(row=row, column=1, value=cat_name).border = s["border"]

        c = ws.cell(row=row, column=2, value=f"=COUNTA('{sname}'!A2:A{last_data_row})")
        c.border = s["border"]
        c.alignment = s["center"]
        count_cells.append(f"B{row}")

        c = ws.cell(row=row, column=3, value=f"='{sname}'!D{last_data_row + 1}")
        c.border = s["border"]
        c.alignment = s["right"]
        c.number_format = '#,##0.00'
        amount_cells.append(f"C{row}")

        ws.cell(row=row, column=4, value="").border = s["border"]
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="合计").font = s["header_font"]
    ws.cell(row=row, column=1).border = s["border"]

    c = ws.cell(row=row, column=2, value=f"=SUM({','.join(count_cells)})" if count_cells else 0)
    c.font = s["header_font"]
    c.border = s["border"]
    c.alignment = s["center"]

    c = ws.cell(row=row, column=3, value=f"=SUM({','.join(amount_cells)})" if amount_cells else 0)
    c.font = s["header_font"]
    c.border = s["border"]
    c.alignment = s["right"]
    c.number_format = '#,##0.00'

    ws.cell(row=row, column=4, value="").border = s["border"]

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

    report_path = output_path / "报销统计.xlsx"
    wb.save(str(report_path))
    print(f"报表已生成: {report_path}")
    return str(report_path)


def generate_summary_report(output_dir: str, periods: list) -> str:
    """生成多周期总汇总报表

    periods 格式:
    [
      {
        "period_name": "0106-0108北京出差",
        "invoice_count": 10,
        "total_amount": 1580.50,
        "categories": {
          "打车票": {"count": 3, "amount": 185.50},
          "火车飞机票": {"count": 2, "amount": 526.00},
          ...
        }
      }
    ]
    """
    output_path = Path(output_dir)
    wb = Workbook()
    s = get_styles()
    category_order = ['打车票', '火车飞机票', '住宿费', '餐费', '其他']

    # --- Sheet 1: 总汇总 ---
    ws = wb.active
    ws.title = "总汇总"

    ws.merge_cells('A1:D1')
    ws['A1'] = f"报销总汇总表 - {datetime.now().strftime('%Y-%m-%d')}"
    ws['A1'].font = s["title_font"]
    ws['A1'].alignment = s["center"]

    headers = ['报销周期', '发票数量', '金额（元）', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = s["header_font_white"]
        cell.fill = s["header_fill"]
        cell.border = s["border"]
        cell.alignment = s["center"]

    row = 4
    for period in periods:
        ws.cell(row=row, column=1, value=period['period_name']).border = s["border"]

        c = ws.cell(row=row, column=2, value=period.get('invoice_count', 0))
        c.border = s["border"]
        c.alignment = s["center"]

        c = ws.cell(row=row, column=3, value=period.get('total_amount', 0))
        c.border = s["border"]
        c.alignment = s["right"]
        c.number_format = '#,##0.00'

        ws.cell(row=row, column=4, value="").border = s["border"]
        row += 1

    # Total
    ws.cell(row=row, column=1, value="合计").font = s["header_font"]
    ws.cell(row=row, column=1).border = s["border"]

    c = ws.cell(row=row, column=2, value=f"=SUM(B4:B{row - 1})")
    c.font = s["header_font"]
    c.border = s["border"]
    c.alignment = s["center"]

    c = ws.cell(row=row, column=3, value=f"=SUM(C4:C{row - 1})")
    c.font = s["header_font"]
    c.border = s["border"]
    c.alignment = s["right"]
    c.number_format = '#,##0.00'

    ws.cell(row=row, column=4, value="").border = s["border"]

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

    # --- Sheet 2: 分类明细 ---
    ws2 = wb.create_sheet("分类明细")

    ws2.merge_cells('A1:F1')
    ws2['A1'] = "各周期分类明细"
    ws2['A1'].font = s["title_font"]
    ws2['A1'].alignment = s["center"]

    headers2 = ['报销周期', '打车票', '火车飞机票', '住宿费', '餐费', '其他']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = s["header_font_white"]
        cell.fill = s["header_fill"]
        cell.border = s["border"]
        cell.alignment = s["center"]

    row = 4
    for period in periods:
        ws2.cell(row=row, column=1, value=period['period_name']).border = s["border"]
        cats = period.get('categories', {})
        for col_idx, cat_name in enumerate(category_order, 2):
            cat_data = cats.get(cat_name, {})
            count = cat_data.get('count', 0)
            amount = cat_data.get('amount', 0)
            if count > 0:
                display = f"{count}张 ¥{amount:,.2f}"
            else:
                display = "-"
            c = ws2.cell(row=row, column=col_idx, value=display)
            c.border = s["border"]
            c.alignment = s["center"]
        row += 1

    ws2.column_dimensions['A'].width = 25
    for col in range(2, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    report_path = output_path / "汇总报销统计.xlsx"
    wb.save(str(report_path))
    print(f"总汇总报表已生成: {report_path}")
    return str(report_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(f"用法:")
        print(f"  python {sys.argv[0]} <output_dir> <invoice_data.json>")
        print(f"  python {sys.argv[0]} <output_dir> <periods_summary.json> --summary")
        sys.exit(1)

    output_dir = sys.argv[1]
    json_path = sys.argv[2]
    is_summary = "--summary" in sys.argv

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if is_summary:
        generate_summary_report(output_dir, data)
    else:
        generate_report(output_dir, data)
